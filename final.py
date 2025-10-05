import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import re


class GradeAnalyzerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("成绩分层统计生成器")
        self.root.geometry("620x420")
        self.root.resizable(False, False)

        self.file_path = None
        self.sheet_names = []
        self.selected_sheet = None

        # ✅ 默认学科列表（可自定义）
        self.default_subjects = ["语文", "数学", "英语", "物理", "化学", "生物", "政治"]
        self.custom_subjects = self.default_subjects.copy()
        self.subject_thresholds = {
            "语文": [300, 400, 600, 700, 800, 900],
            "数学": [300, 400, 600, 700, 800, 900],
            "英语": [300, 400, 600, 700, 800, 900],
            "物理": [300, 400, 600, 700],
            "化学": [300, 400, 600],
            "生物": [300, 400, 500],
            "政治": [100, 200, 300],
        }

        tk.Label(root, text="成绩分层统计生成器", font=("微软雅黑", 14, "bold")).pack(pady=10)

        tk.Button(root, text="📁 选择成绩文件 (.xlsx)", command=self.select_file, font=("微软雅黑", 12), width=30).pack(pady=5)
        self.file_label = tk.Label(root, text="未选择文件", fg="gray", font=("微软雅黑", 10))
        self.file_label.pack()

        sheet_frame = tk.Frame(root)
        sheet_frame.pack(pady=5)
        tk.Label(sheet_frame, text="选择工作表：", font=("微软雅黑", 10)).pack(side="left")
        self.sheet_combo = ttk.Combobox(sheet_frame, state="readonly", width=30)
        self.sheet_combo.pack(side="left", padx=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)

        # 替换原来的“自定义阈值”按钮为“学科与阈值设置”
        tk.Button(root, text="📚 学科与阈值设置", command=self.open_subject_settings, font=("微软雅黑", 11), width=30).pack(pady=10)

        self.start_btn = tk.Button(root, text="🚀 开始生成分层统计表", command=self.process_file, font=("微软雅黑", 12), width=30, state="disabled")
        self.start_btn.pack(pady=15)

        self.progress = ttk.Progressbar(root, orient="horizontal", length=450, mode="determinate")
        self.progress.pack(pady=5)

        self.status_label = tk.Label(root, text="", fg="blue", font=("微软雅黑", 9))
        self.status_label.pack()

        tk.Label(
            root,
            text="提示：原始表格中总分排名列请命名为“总分校方向排名”，各科排名列请命名为“XX校方向名次”或“XX校次”（如：语文校方向名次）",
            font=("微软雅黑", 10),
            fg="red",
            wraplength=600,
            justify="left"
        ).pack(side="bottom", pady=(10, 5))

    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="选择成绩 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            self.file_path = file_path
            self.file_label.config(text=os.path.basename(file_path), fg="black")

            try:
                excel_file = pd.ExcelFile(file_path)
                self.sheet_names = excel_file.sheet_names
                excel_file.close()

                if not self.sheet_names:
                    raise ValueError("该 Excel 文件中没有工作表")

                self.sheet_combo['values'] = self.sheet_names
                self.sheet_combo.set("")
                self.selected_sheet = None
                self.start_btn.config(state="disabled")

            except Exception as e:
                messagebox.showerror("错误", f"无法读取 Excel 文件的 sheet 列表：\n{str(e)}")
                self.file_path = None
                self.file_label.config(text="未选择文件", fg="gray")

    def on_sheet_selected(self, event=None):
        selected = self.sheet_combo.get()
        if selected in self.sheet_names:
            self.selected_sheet = selected
            self.start_btn.config(state="normal")
        else:
            self.selected_sheet = None
            self.start_btn.config(state="disabled")

    def parse_threshold_input(self, text):
        if not text.strip():
            return []
        parts = re.split(r'[，,\s]+', text.strip())
        try:
            return [int(x) for x in parts if x.isdigit()]
        except ValueError:
            return None

    def open_subject_settings(self):
        top = tk.Toplevel(self.root)
        top.title("学科与阈值设置")
        top.geometry("500x500")
        top.transient(self.root)
        top.grab_set()

        tk.Label(top, text="自定义学科及其分层阈值", font=("微软雅黑", 12, "bold")).pack(pady=5)

        # 学科列表框
        list_frame = tk.Frame(top)
        list_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.subject_listbox = tk.Listbox(list_frame, width=20, height=12)
        self.subject_listbox.pack(side="left", fill="y")

        for subj in self.custom_subjects:
            self.subject_listbox.insert(tk.END, subj)

        btn_frame = tk.Frame(list_frame)
        btn_frame.pack(side="left", padx=10)

        def add_subject():
            name = tk.simpledialog.askstring("添加学科", "请输入学科名称（如：技术、美术等）：")
            if name and name.strip():
                name = name.strip()
                if name not in self.custom_subjects:
                    self.custom_subjects.append(name)
                    self.subject_thresholds[name] = []
                    self.subject_listbox.insert(tk.END, name)
                else:
                    messagebox.showwarning("提示", f"学科“{name}”已存在！")

        def remove_subject():
            sel = self.subject_listbox.curselection()
            if sel:
                idx = sel[0]
                subj = self.custom_subjects[idx]
                del self.custom_subjects[idx]
                del self.subject_thresholds[subj]
                self.subject_listbox.delete(idx)
            else:
                messagebox.showwarning("提示", "请先选择一个学科")

        tk.Button(btn_frame, text="➕ 添加", command=add_subject, width=10).pack(pady=5)
        tk.Button(btn_frame, text="➖ 删除", command=remove_subject, width=10).pack(pady=5)

        # 阈值输入区
        input_frame = tk.Frame(top)
        input_frame.pack(pady=10, padx=10, fill="x")

        tk.Label(input_frame, text="当前学科阈值：", anchor="w").pack(anchor="w")
        self.threshold_var = tk.StringVar()
        threshold_entry = tk.Entry(input_frame, textvariable=self.threshold_var, font=("微软雅黑", 10), width=50)
        threshold_entry.pack(fill="x", pady=5)

        def on_subject_select(event):
            sel = self.subject_listbox.curselection()
            if sel:
                subj = self.custom_subjects[sel[0]]
                vals = self.subject_thresholds.get(subj, [])
                self.threshold_var.set(",".join(map(str, vals)))

        def save_threshold():
            sel = self.subject_listbox.curselection()
            if not sel:
                messagebox.showwarning("提示", "请先选择一个学科")
                return
            subj = self.custom_subjects[sel[0]]
            val = self.parse_threshold_input(self.threshold_var.get())
            if val is None:
                messagebox.showerror("错误", "阈值格式不正确！请用数字和逗号/空格分隔。")
                return
            self.subject_thresholds[subj] = sorted(val)
            messagebox.showinfo("成功", f"“{subj}”的阈值已保存！")

        self.subject_listbox.bind("<<ListboxSelect>>", on_subject_select)
        tk.Button(input_frame, text="💾 保存当前学科阈值", command=save_threshold, bg="#4CAF50", fg="white").pack(pady=5)

        # 底部按钮
        def restore_defaults():
            self.custom_subjects = self.default_subjects.copy()
            self.subject_thresholds = {
                "语文": [300, 400, 600, 700, 800, 900],
                "数学": [300, 400, 600, 700, 800, 900],
                "英语": [300, 400, 600, 700, 800, 900],
                "物理": [300, 400, 600, 700],
                "化学": [300, 400, 600],
                "生物": [300, 400, 500],
                "政治": [100, 200, 300],
                "历史": [100, 200, 300],
                "地理": [100, 200, 300],
            }
            self.subject_listbox.delete(0, tk.END)
            for subj in self.custom_subjects:
                self.subject_listbox.insert(tk.END, subj)
            self.threshold_var.set("")

        def close_window():
            top.destroy()

        bottom_frame = tk.Frame(top)
        bottom_frame.pack(pady=10)
        tk.Button(bottom_frame, text="恢复默认", command=restore_defaults, font=("微软雅黑", 10)).pack(side="left", padx=5)
        tk.Button(bottom_frame, text="完成", command=close_window, font=("微软雅黑", 10), bg="#2196F3", fg="white").pack(side="left", padx=5)

    def process_file(self):
        if not self.file_path or not self.selected_sheet:
            messagebox.showwarning("警告", "请先选择文件和工作表！")
            return

        try:
            self.status_label.config(text="正在读取文件...")
            self.root.update()

            df = pd.read_excel(self.file_path, sheet_name=self.selected_sheet)

            total_rank_col = None
            candidate_cols = [
                "总分校方向名次",
                "总分校次",
                "总分排名",
                "总分名次",
                "总分 校次"
            ]
            for col in candidate_cols:
                if col in df.columns:
                    total_rank_col = col
                    break

            if total_rank_col is None:
                raise KeyError("未找到总分排名列。请确保表中包含类似“总分校次”的列。")

            df = df[~df[total_rank_col].astype(str).str.contains("不计排名", na=False)]
            df[total_rank_col] = pd.to_numeric(df[total_rank_col], errors='coerce')
            df = df.dropna(subset=[total_rank_col])

            def find_rank_col(subject):
                for col in df.columns:
                    if subject in col and ('校次' in col or '排名' in col or '校方向名次' in col):
                        return col
                return None

            # ✅ 只处理用户自定义的学科
            subjects_info = {}
            for subj in self.custom_subjects:
                score_col = subj
                rank_col = find_rank_col(subj)
                thresholds = self.subject_thresholds.get(subj, [])
                if score_col in df.columns and rank_col is not None:
                    subjects_info[subj] = {
                        "score_col": score_col,
                        "rank_col": rank_col,
                        "thresholds": thresholds
                    }

            if not subjects_info:
                raise ValueError("未找到任何有效的科目数据（请检查学科名称是否与 Excel 列名一致）")

            for info in subjects_info.values():
                rank_col = info["rank_col"]
                df[rank_col] = pd.to_numeric(df[rank_col], errors='coerce')

            dir_name = os.path.dirname(self.file_path)
            base_name = os.path.splitext(os.path.basename(self.file_path))[0]
            output_file = os.path.join(dir_name, f"{base_name}_分层统计表.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "分层统计汇总"

            start_col = 1
            total_subjects = len(subjects_info)
            for idx, (subj_name, info) in enumerate(subjects_info.items(), 1):
                self.status_label.config(text=f"处理中：{subj_name} ({idx}/{total_subjects})")
                self.progress['value'] = (idx / total_subjects) * 100
                self.root.update()

                score_col = info["score_col"]
                rank_col = info["rank_col"]
                thresholds = info["thresholds"]

                sub_df = df[["姓名", score_col, rank_col]].copy()
                sub_df = sub_df.dropna(subset=[rank_col])
                sub_df = sub_df.sort_values(by=rank_col).reset_index(drop=True)

                ws.cell(row=1, column=start_col, value="姓名")
                ws.cell(row=1, column=start_col + 1, value=f"{subj_name}")
                ws.cell(row=1, column=start_col + 2, value=f"{subj_name}校次")

                current_row = 2
                threshold_idx = 0
                n = len(sub_df)

                for i in range(n):
                    current_rank = sub_df.iloc[i][rank_col]
                    while threshold_idx < len(thresholds) and current_rank > thresholds[threshold_idx]:
                        ws.cell(row=current_row, column=start_col, value=f"前{thresholds[threshold_idx]}名 共{i}人")
                        current_row += 1
                        threshold_idx += 1

                    ws.cell(row=current_row, column=start_col, value=sub_df.iloc[i]["姓名"])
                    ws.cell(row=current_row, column=start_col + 1, value=sub_df.iloc[i][score_col])
                    ws.cell(row=current_row, column=start_col + 2, value=int(sub_df.iloc[i][rank_col]))
                    current_row += 1

                while threshold_idx < len(thresholds):
                    ws.cell(row=current_row, column=start_col, value=f"前{thresholds[threshold_idx]}名 共{n}人")
                    current_row += 1
                    threshold_idx += 1

                if thresholds:
                    last_th = thresholds[-1]
                    after_count = (sub_df[rank_col] > last_th).sum()
                    if after_count > 0:
                        ws.cell(row=current_row, column=start_col, value=f"{last_th}名以后 共{int(after_count)}人")
                        current_row += 1

                for r in range(2, current_row):
                    cell_val = ws.cell(row=r, column=start_col).value
                    if isinstance(cell_val, str) and ("前" in cell_val or "以后" in cell_val):
                        ws.cell(row=r, column=start_col).font = Font(bold=True)

                start_col += 4

            wb.save(output_file)
            self.status_label.config(text="✅ 处理完成！")
            messagebox.showinfo("成功", f"合并版分层统计表已生成：\n{output_file}")

        except Exception as e:
            self.status_label.config(text="❌ 处理失败")
            messagebox.showerror("错误", f"发生错误：\n{str(e)}")
        finally:
            self.progress['value'] = 0


if __name__ == "__main__":
    root = tk.Tk()
    app = GradeAnalyzerGUI(root)
    root.mainloop()